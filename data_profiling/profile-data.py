import argparse
from collections import Counter, defaultdict
import csv
from datetime import datetime
from pathlib import Path
import os
import random
import re
import sys
import tempfile
import zipfile
# Imports above are standard Python
# Imports below are 3rd-party
from lib.base import C, Database, Logger, get_line_count
from argparse_range import range_action
from dateutil.parser import parse
from dotenv import dotenv_values
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, borders
import pandas as pd
import seaborn as sns

MAX_SHEET_NAME_LENGTH = 31  # Excel limitation
ROUNDING = 1  # 5.4% for example
VALUE, COUNT = "Value", "Count"
NUMBER, DATETIME, STRING = "NUMBER", "DATETIME", "STRING"

DATATYPE_MAPPING_DICT = {
    "BIGINT": NUMBER,
    "BINARY": NUMBER,
    "BIT": NUMBER,
    "BOOLEAN": NUMBER,
    "DECIMAL": NUMBER,
    "DOUBLE": NUMBER,
    "FLOAT": NUMBER,
    "INTEGER": NUMBER,
    "NUMERIC": NUMBER,
    "REAL": NUMBER,
    "SMALLINT": NUMBER,
    "TINYINT": NUMBER,
    "VARBINARY": NUMBER,

    "DATE": DATETIME,
    "TIMESTAMP": DATETIME,

    "BLOB": STRING,
    "CHAR": STRING,
    "CLOB": STRING,
    "LONGNVARCHAR": STRING,
    "LONGVARBINARY": STRING,
    "LONGVARCHAR": STRING,
    "NCHAR": STRING,
    "NCLOB": STRING,
    "NVARCHAR": STRING,
    "OTHER": STRING,
    "SQLXML": STRING,
    "TIME": STRING,
    "VARCHAR": STRING,
}

# When producing a list of detail values and their frequency of occurrence
DEFAULT_MAX_DETAIL_VALUES = 35
# When analyzing the patterns of a string column
DEFAULT_MAX_PATTERN_LENGTH = 50
# Don't plot distributions if there are fewer than this number of distinct values
DISTRIBUTION_PLOT_MIN_VALUES = 6
# Categorical plots should have no more than this number of distinct values
# Groups the rest in "Other"
CATEGORICAL_PLOT_MAX_VALUES = 5
OTHER = "Other"

# Plotting visual effects
PLOT_SIZE_X, PLOT_SIZE_Y = 11, 8.5
PLOT_FONT_SCALE = 0.75

ROW_COUNT = "count"
NULL_COUNT = "null"
NULL_PERCENT = "%null"
UNIQUE_COUNT = "unique"
UNIQUE_PERCENT = "%unique"
MOST_COMMON = "most_common"
MOST_COMMON_PERCENT = "%most_common"
LARGEST = "largest"
SMALLEST = "smallest"
LONGEST = "longest"
SHORTEST = "shortest"
MEAN = "mean"
PERCENTILE_25TH = "percentile_25th"
MEDIAN = "median"
PERCENTILE_75TH = "percentile_75th"
STDDEV = "stddev"

ANALYSIS_LIST = (
    ROW_COUNT,
    NULL_COUNT,
    NULL_PERCENT,
    UNIQUE_COUNT,
    UNIQUE_PERCENT,
    MOST_COMMON,
    MOST_COMMON_PERCENT,
    LARGEST,
    SMALLEST,
    LONGEST,
    SHORTEST,
    MEAN,
    PERCENTILE_25TH,
    MEDIAN,
    PERCENTILE_75TH,
    STDDEV,
)


def truncate_string(s: str, max_length: int, filler: str = "...") -> str:
    """
    For example, truncate_string("Hello world!", 7) returns:
    "Hell..."
    """
    excess_count = len(s) - max_length
    if excess_count <= 0:
        return s
    else:
        return s[:max_length - len(filler)] + filler


def get_pattern(l: list) -> dict:
    """
    Return a Counter where the keys are the observed patterns and the values are how often they appear.
    Examples:
    "hi joe." --> "CC_CCC."
    "hello4abigail" --> "C(5)9C(7)"
    :param l: a list of strings
    :return: a pattern analysis
    """
    counter = Counter()
    for value in l:
        if not value:
            continue
        value = re.sub("[a-zA-Z]", "C", value)  # Replace letters with 'C'
        value = re.sub(r"\d", "9", value)  # Replace numbers with '9'
        value = re.sub(r"\s+", "_", value)  # Replace whitespace with '_'
        value = re.sub(r"\W", "?", value)  # Replace whitespace with '?'
        # Group long sequences of letters or numbers
        # See https://stackoverflow.com/questions/76230795/replace-characters-with-a-count-of-characters
        # The number below (2) means sequences of 3 or more will be grouped
        value = re.sub(r'(.)\1{2,}', lambda m: f'{m.group(1)}({len(m.group())})', value)
        counter[value] += 1
    return counter


def parse_date(a_date: str) -> datetime:
    try:
        return parse(a_date)  # dateutil's parser
    except:
        return None


def select_random_rows(input_file, output_file, n):
    """
    Selects random n rows from a CSV file and exports them to a new CSV file.

    Args:
        input_file (str): Path to the input CSV file.
        output_file (str): Path to the output CSV file.
        n (int): Number of random rows to select.

    Returns:
        None
    """
    # Read the input CSV file into a pandas DataFrame
    df = pd.read_csv(input_file)

    # Check if the number of rows in the DataFrame is less than n
    if len(df) < n:
        raise ValueError("The number of rows in the input file is less than n.")

    # Select random n rows from the DataFrame
    random_rows = df.sample(n)

    # Export the selected random rows to a new CSV file
    random_rows.to_csv(output_file, index=False)


parser = argparse.ArgumentParser(
    description='Profile the data in a database or CSV file.',
    epilog='Generates an analysis consisting of an Excel workbook and (optionally) one or more images.'
)
parser.add_argument('input',
                    metavar="/path/to/input_data_file.csv | query",
                    help="If a file no connection information required.")
parser.add_argument('--db-host-name',
                    metavar="HOST_NAME",
                    help="Overrides HOST_NAME environment variable.")
parser.add_argument('--db-port-number',
                    metavar="PORT_NUMBER",
                    help="Overrides PORT_NUMBER environment variable.")
parser.add_argument('--db-name',
                    metavar="DATABASE_NAME",
                    help="Overrides DATABASE_NAME environment variable.")
parser.add_argument('--db-user-name',
                    metavar="USER_NAME",
                    help="Overrides USER_NAME environment variable.")
parser.add_argument('--db-password',
                    metavar="PASSWORD",
                    help="Overrides PASSWORD environment variable.")
parser.add_argument('--environment-file',
                    metavar="/path/to/file",
                    help="An additional source of database connection information. Overrides environment settings.")
parser.add_argument('--header-lines',
                    type=int,
                    metavar="NUM",
                    action=range_action(1, sys.maxsize),
                    default=0,
                    help="When reading from a file specifies the number of rows to skip for header information. Ignored when getting data from a database. Default is 0.")
parser.add_argument('--sample-rows-file',
                    type=int,
                    metavar="NUM",
                    action=range_action(1, sys.maxsize),
                    help=f"When reading from a file randomly choose this number of rows. If greater than or equal to the number of data rows will use all rows. Ignored when getting data from a database.")
parser.add_argument('--max-detail-values',
                    type=int,
                    metavar="NUM",
                    action=range_action(1, sys.maxsize),
                    default=DEFAULT_MAX_DETAIL_VALUES,
                    help=f"Produce this many of the top/bottom value occurrences, default is {DEFAULT_MAX_DETAIL_VALUES}.")
parser.add_argument('--max-pattern-length',
                    type=int,
                    metavar="NUM",
                    action=range_action(1, sys.maxsize),
                    default=DEFAULT_MAX_PATTERN_LENGTH,
                    help=f"When segregating strings into patterns leave untouched strings of length greater than this, default is {DEFAULT_MAX_PATTERN_LENGTH}.")
parser.add_argument('--output-dir',
                    metavar="/path/to/dir",
                    default=Path.cwd(),
                    help="Default is the current directory.")

logging_group = parser.add_mutually_exclusive_group()
logging_group.add_argument('-v', '--verbose', action='store_true')
logging_group.add_argument('-t', '--terse', action='store_true')

args = parser.parse_args()
if args.input.endswith(C.CSV_EXTENSION):
    input_path = Path(args.input)
    input_query = None
else:
    input_path = None
    input_query = args.input
host_name = args.db_host_name
port_number = args.db_port_number
database_name = args.db_name
user_name = args.db_user_name
password = args.db_password
if args.environment_file:
    environment_file = Path(args.environment_file)
else:
    environment_file = ""
header_lines = args.header_lines
sample_rows_file = args.sample_rows_file
max_detail_values = args.max_detail_values
max_pattern_length = args.max_pattern_length
output_dir = Path(args.output_dir)

environment_settings_dict = {
    **os.environ,
    **dotenv_values(environment_file),
}
if not output_dir.exists():
    parser.error("Directory '{output_dir}' does not exist.")
if input_query:
    # Verify we have the information we need to connect to the database
    host_name = host_name or environment_settings_dict["HOST_NAME"]
    port_number = port_number or environment_settings_dict["PORT_NUMBER"]
    database_name = database_name or environment_settings_dict["DATABASE_NAME"]
    user_name = user_name or environment_settings_dict["USER_NAME"]
    password = password or environment_settings_dict["PASSWORD"]
    if not host_name and port_number and database_name and user_name and password:
        parser.error("Connecting to a database requires: --db-host-name, --db-port-number, --db-name, --db-user-name, --db-password")
elif input_path:
    if not input_path.exists():
        parser.error(f"Could not find input file '{input_path}'.")
else:
    raise Exception("Programming error.")

if args.verbose:
    logger = Logger().get_logger("DEBUG")
elif args.terse:
    logger = Logger().get_logger("WARNING")
else:
    logger = Logger().get_logger()

# Now, read the data
data_dict = defaultdict(list)
datetype_dict = None
if input_query:
    # User wants to get the data from a database query
    mydb = Database(
        host_name=host_name,
        port_number=port_number,
        database_name=database_name,
        user_name=user_name,
        password=password
    )
    cursor, column_list = mydb.execute(input_query)
    for r in cursor.fetchall():
        row = dict(zip(column_list, r))
        for column_name, value in row.items():
            data_dict[column_name].append(value)
elif input_path:
    # Data is coming from a file
    logger.info(f"Reading from '{input_path}' ...")
    types = defaultdict(str, A="str")  # Pandas is pretty terrible at determining datatypes, but using it for ingestion because it is fast
    df = pd.read_csv(input_path, dtype='object')
    if sample_rows_file:
        df = df.sample(sample_rows_file)
    data_dict = df.to_dict(orient='list')
    # Set best type for each column of data
    for column_name, values in data_dict.items():
        # Sample up to 100 non-null values
        non_null_list = [x for x in values if x]
        sampled_list = random.sample(non_null_list, min(100, len(non_null_list)))
        is_parse_error = False
        for item in non_null_list:
            try:
                parse_date(item)
            except parser._parser.ParserError:
                is_parse_error = True
                logger.info(f"Cannot cast column '{column_name}' as a datetime.")
                break
        if not is_parse_error:
            logger.info(f"Casting column '{column_name}' as a datetime.")
            data_dict[column_name] = map(lambda x: parse(x), values)
        else:
            # Not a datetime, try number
            is_parse_error = False
            for item in non_null_list:
                try:
                    float(item)
                except ValueError:
                    is_parse_error = True
                    logger.info(f"Cannot cast column '{column_name}' as a number.")
                    break
            if not is_parse_error:
                logger.info(f"Casting column '{column_name}' as a number.")
                data_dict[column_name] = map(lambda x: float(x), values)
            else:
                pass  # Currently str values and that's what we will leave them as

# Data has been read into input_df, now process it
# To temporarily hold distribution plots
tempdir = tempfile.TemporaryDirectory()
tempdir_path = Path(tempdir.name)
# To keep track of which columns have distribution plots
distribution_plot_list = list()

summary_dict = dict()  # To be converted into the summary worksheet
detail_dict = dict()  # Each element to be converted into a detail worksheet
pattern_dict = dict()  # For each string column calculate the frequency of patterns
for column_name in input_df.columns:
    logger.info(f"Working on column '{column_name}' ...")
    input_df[column_name] = set_best_type(input_df[column_name])
    data = input_df[column_name]
    logger.debug(f"Treating this column as data type '{data.dtype}'.")
    row_dict = dict.fromkeys(ANALYSIS_LIST)
    # Row count
    row_count = data.size
    if not row_count:
        logger.warning("No data.")
        sys.exit()
    row_dict[ROW_COUNT] = row_count
    # Null
    null_count = row_count - data.count()
    row_dict[NULL_COUNT] = null_count
    # Null%
    row_dict[NULL_PERCENT] = round(100 * null_count / row_count, ROUNDING)
    # Unique
    unique_count = len(data.unique())
    row_dict[UNIQUE_COUNT] = unique_count
    # Unique%
    row_dict[UNIQUE_PERCENT] = round(100 * unique_count / row_count, ROUNDING)

    if null_count != row_count:
        # Most common (mode)
        row_dict[MOST_COMMON] = list(data.mode().values)[0]
        # Most common%
        row_dict[MOST_COMMON_PERCENT] = round(100 * list(data.value_counts())[0] / row_count, ROUNDING)

        if data.dtype == OBJECT:
            # Largest & smallest
            row_dict[LARGEST] = data.dropna().astype(pd.StringDtype()).max()
            row_dict[SMALLEST] = data.dropna().astype(pd.StringDtype()).min()
            # Longest & shortest
            row_dict[LONGEST] = max(data.dropna().astype(pd.StringDtype()).values, key=len)
            row_dict[SHORTEST] = min(data.dropna().astype(pd.StringDtype()).values, key=len)
            # No mean/quartiles/stddev statistics for strings
        else:  # numeric or datetime
            # Largest & smallest
            row_dict[LARGEST] = data.max()
            row_dict[SMALLEST] = data.min()
            # No longest/shortest for dates and numbers
            # Mean/quartiles/stddev statistics
            row_dict[MEAN] = data.mean()
            row_dict[PERCENTILE_25TH] = data.quantile(0.25)
            row_dict[MEDIAN] = data.quantile(0.5)
            row_dict[PERCENTILE_75TH] = data.quantile(0.75)
            row_dict[STDDEV] = data.std()

        # Value counts
        # Collect no more than number of values available or what was given on the command-line
        # whichever is less
        detail_df = pd.DataFrame()
        max_length = min(max_detail_values, len(data.value_counts(dropna=False)))
        # Create 3-column ascending visual
        detail_df["rank"] = list(range(1, max_length + 1))
        detail_df["value"] = list(data.value_counts(dropna=False).index)[:max_length]
        detail_df["count"] = list(data.value_counts(dropna=False))[:max_length]
        percent_total_list = list(data.value_counts(dropna=False, normalize=True))[:max_length]
        detail_df["%total"] = [round(x*100, ROUNDING) for x in percent_total_list]
    else:
        logger.info(f"Column is empty.")

    summary_dict[column_name] = row_dict
    detail_dict[column_name] = detail_df

    # For string columns produce a pattern analysis
    # For numeric and datetime columns produce a distribution plot
    if data.dtype == OBJECT:  # string data
        pattern_df = pd.DataFrame()
        pattern_data = data.apply(get_pattern)
        pattern_analysis = pattern_data.value_counts(normalize=True)
        max_length = min(max_detail_values, len(pattern_data.value_counts(dropna=False)))
        pattern_df["rank"] = list(range(1, max_length + 1))
        pattern_df["pattern"] = list(pattern_data.value_counts(dropna=False).index)[:max_length]
        pattern_df["count"] = list(pattern_data.value_counts(dropna=False))[:max_length]
        percent_total_list = list(pattern_data.value_counts(dropna=False, normalize=True))[:max_length]
        pattern_df["%total"] = [round(x*100, ROUNDING) for x in percent_total_list]
        pattern_dict[column_name] = pattern_df
    else:  # Numeric/datetime data
        sns.set_theme()
        sns.set(font_scale=PLOT_FONT_SCALE)
        plot_data = data.value_counts(normalize=True)
        if len(plot_data) >= DISTRIBUTION_PLOT_MIN_VALUES:
            logger.debug("Creating a distribution plot ...")
            g = sns.displot(data)
            plot_output_path = tempdir_path / f"{column_name}.distribution.png"
            g.set_axis_labels(VALUE, COUNT, labelpad=10)
            g.figure.set_size_inches(PLOT_SIZE_X, PLOT_SIZE_Y)
            g.ax.margins(.15)
            g.savefig(plot_output_path)
            logger.info(f"Wrote {os.stat(plot_output_path).st_size} bytes to '{plot_output_path}'.")
            distribution_plot_list.append(column_name)
        else:
            logger.debug("Not enough distinct values to create a distribution plot.")

# Convert the summary_dict dictionary of dictionaries to a DataFrame
result_df = pd.DataFrame.from_dict(summary_dict, orient='index')
# And write it to a worksheet
logger.info("Writing summary ...")
# Set target filename based on database v. file and whether we are sampling
if host_name and sample_percent:
    output_file = (output_dir / f"{input_path}.sample{sample_percent}pct{EXCEL_EXTENSION}")
elif host_name and not sample_percent:
    output_file = (output_dir / f"{input_path}{EXCEL_EXTENSION}")
elif not host_name and sample_percent:
    output_file = (output_dir / f"{input_path.stem}.sample{sample_percent}pct{EXCEL_EXTENSION}")
elif not host_name and not sample_percent:
    output_file = (output_dir / f"{input_path.stem}{EXCEL_EXTENSION}")
else:
    raise("Programming error.")
writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
result_df.to_excel(writer, sheet_name="Summary")
# And generate a detail sheet, and optionally a pattern sheet, for each column
for column_name, detail_df in detail_dict.items():
    logger.info(f"Writing detail for column '{column_name}' ...")
    detail_df.to_excel(writer, index=False, sheet_name=truncate_string(column_name + " detail", MAX_SHEET_NAME_LENGTH))
    if column_name in pattern_dict:
        logger.info(f"Writing pattern detail for string column '{column_name}' ...")
        pattern_df = pattern_dict[column_name]
        pattern_df.to_excel(writer, index=False, sheet_name=truncate_string(column_name + " pattern", MAX_SHEET_NAME_LENGTH))
writer.close()

# Add the plots and size bars to the Excel file
workbook = openpyxl.load_workbook(output_file)

# Plots
# Look for sheet names corresponding to the plot filename
sheet_number = -1
for sheet_name in workbook.sheetnames:
    sheet_number += 1
    if sheet_number == 0:  # Skip summary sheet (first sheet, zero-based-index)
        continue
    column_name = sheet_name[:-7]  # remove " detail" from sheet name to get column name
    if column_name in distribution_plot_list:
        workbook.create_sheet(column_name + " distribution", sheet_number+1)
        worksheet = workbook.worksheets[sheet_number+1]
        image_path = tempdir_path / (column_name + ".distribution.png")
        logger.info(f"Adding {image_path} to {output_file} after sheet {sheet_name} ...")
        image = openpyxl.drawing.image.Image(image_path)
        image.anchor = "A1"
        worksheet.add_image(image)
        sheet_number += 1

# Size bar columns for the worksheets showing the ranks of values and patterns
for i, sheet_name in enumerate(workbook.sheetnames):
    if sheet_name.endswith("detail") or sheet_name.endswith("pattern"):
        worksheet = workbook.worksheets[i]
        for row_number in range(2, worksheet.max_row+1):
            value_to_convert = worksheet[f'D{row_number}'].value  # C = 3rd column, convert from percentage
            bar_representation = "█" * round(value_to_convert)
            worksheet[f'E{row_number}'] = bar_representation
        # And set some visual formatting while we are here
        worksheet.column_dimensions['B'].width = 25
        # worksheet.column_dimensions['C'].number_format = "0.0"

# Formatting for the summary sheet
worksheet = workbook.worksheets[0]
worksheet.column_dimensions['A'].width = 25  # Column names
worksheet.column_dimensions['G'].width = 15  # Most common value
worksheet.column_dimensions['I'].width = 15  # Largest value
worksheet.column_dimensions['J'].width = 15  # Smallest value
worksheet.column_dimensions['K'].width = 15  # Longest value

for row in range(1, worksheet.max_row+1):
    worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='right')
for row in range(1, worksheet.max_row+1):
    for col in range(1, 17):
        worksheet.cell(row=row, column=col).border = Border(outline=Side(border_style=borders.BORDER_THICK, color='FFFFFFFF'))

workbook.save(output_file)
logger.info(f"Wrote {os.stat(output_file).st_size} bytes to '{output_file}'.")
